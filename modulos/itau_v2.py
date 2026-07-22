# -*- coding: utf-8 -*-
"""
Simulador Itaú V 2.0 — módulo do ERP (Streamlit).

Lê a Guia de Oportunidades MAIS RECENTE direto do Google Drive (via o mesmo
service account que o ERP já usa), converte para JSON e injeta no itau_v2.html.
Sempre reflete a última Guia — sem git push, sem CORS.

Estratégia de busca (à prova de ID errado):
  1) tenta listar a pasta do secret DRIVE_FOLDER_ITAU;
  2) se falhar ou vier vazia, procura a Guia PELO NOME em tudo que o service
     account enxerga (basta a pasta estar compartilhada com ele).

Requer:
  - Secret `gcp_service_account` (já existe no ERP)
  - Secret `DRIVE_FOLDER_ITAU` (opcional, mas recomendado) = ID da pasta Tabelas/ITAU
  - openpyxl no requirements.txt
"""
import os
import io
import re
import json
import datetime

import streamlit as st
import streamlit.components.v1 as components
from googleapiclient.errors import HttpError

from utils import get_drive_service  # reutiliza a autenticação já existente do ERP

_CAMPOS = "files(id, name, mimeType, modifiedTime)"

# ------------------------------------------------------------------ #
# Datas no nome do arquivo
# ------------------------------------------------------------------ #
_MESES = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4, "maio": 5, "junho": 6,
    "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}


def _mk(ano, mes, dia):
    if ano < 100:
        ano += 2000
    try:
        return datetime.date(ano, mes, dia)
    except (ValueError, TypeError):
        return None


def _data_do_nome(nome):
    """Extrai a data do nome (mês por extenso, AAAAMMDD, AAAA-MM-DD, DD-MM-AAAA)."""
    m = re.search(r"(\d{1,2})[-_ .]+(?:de[-_ ]+)?([A-Za-zçÇãÃáàâéêíóôõúÁÉ]{3,})[-_ .]+(?:de[-_ ]+)?(\d{2,4})", nome)
    if m:
        mes = (m.group(2).lower()
               .replace("ç", "c").replace("ã", "a").replace("â", "a").replace("á", "a")
               .replace("à", "a").replace("é", "e").replace("ê", "e").replace("í", "i")
               .replace("ó", "o").replace("ô", "o").replace("õ", "o").replace("ú", "u"))
        if mes in _MESES:
            d = _mk(int(m.group(3)), _MESES[mes], int(m.group(1)))
            if d:
                return d
    for g in re.findall(r"(20\d{2})(\d{2})(\d{2})", nome):
        d = _mk(int(g[0]), int(g[1]), int(g[2]))
        if d:
            return d
    m = re.search(r"(20\d{2})[-_ ./](\d{1,2})[-_ ./](\d{1,2})", nome)
    if m:
        d = _mk(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        if d:
            return d
    m = re.search(r"(\d{1,2})[-_ ./](\d{1,2})[-_ ./](\d{2,4})", nome)
    if m:
        d = _mk(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        if d:
            return d
    return None


# ------------------------------------------------------------------ #
# Google Drive
# ------------------------------------------------------------------ #
def _msg_http(e):
    """Extrai status + mensagem legível de um HttpError do Google."""
    status = getattr(getattr(e, "resp", None), "status", "?")
    detalhe = ""
    try:
        detalhe = json.loads(e.content.decode("utf-8")).get("error", {}).get("message", "")
    except Exception:
        detalhe = str(e)[:300]
    dica = ""
    if status == 404:
        dica = " → ID da pasta errado OU a pasta não está compartilhada com o service account."
    elif status == 403:
        dica = " → Sem permissão na pasta OU a Google Drive API não está ativada no projeto."
    return f"HTTP {status}: {detalhe}{dica}"


def _diagnostico_visibilidade(service):
    """Lista o que o service account realmente enxerga — prova se o compartilhamento funcionou."""
    try:
        res = service.files().list(
            q="trashed=false", fields="files(id, name, mimeType)", pageSize=30,
            supportsAllDrives=True, includeItemsFromAllDrives=True,
        ).execute()
        itens = res.get("files", [])
        if not itens:
            return ("❌ O service account NÃO enxerga NENHUM arquivo ou pasta.\n"
                    "Ou nada foi compartilhado com ele, ou a Google Drive API está desativada no projeto.")
        linhas = []
        for i in itens:
            icone = "📁" if i.get("mimeType") == "application/vnd.google-apps.folder" else "📄"
            linhas.append(f"{icone} {i['name']}   (id: {i['id']})")
        return "✅ O service account enxerga estes itens:\n" + "\n".join(linhas)
    except HttpError as e:
        return f"Falha ao listar o que o service account enxerga → {_msg_http(e)}"


def _so_guias(arquivos):
    return [a for a in arquivos
            if "oportunidad" in a.get("name", "").lower()
            and a.get("name", "").lower().endswith((".xlsx", ".xlsm"))
            and not a.get("name", "").startswith("~$")]


def _buscar_guias(service, folder_id):
    """Retorna (guias, log). Tenta pela pasta; se falhar/vier vazio, busca pelo nome."""
    log = []
    guias = []

    if folder_id:
        try:
            res = service.files().list(
                q=f"'{folder_id}' in parents and trashed=false",
                fields=_CAMPOS, pageSize=200,
                supportsAllDrives=True, includeItemsFromAllDrives=True,
            ).execute()
            arquivos = res.get("files", [])
            guias = _so_guias(arquivos)
            log.append(f"Pasta {folder_id}: {len(arquivos)} arquivo(s), {len(guias)} Guia(s).")
        except HttpError as e:
            log.append(f"Falha ao ler a pasta {folder_id} → {_msg_http(e)}")
    else:
        log.append("Secret DRIVE_FOLDER_ITAU não definido — indo direto para a busca por nome.")

    if not guias:
        try:
            res = service.files().list(
                q="name contains 'oportunidad' and trashed=false",
                fields=_CAMPOS, pageSize=200,
                supportsAllDrives=True, includeItemsFromAllDrives=True,
            ).execute()
            arquivos = res.get("files", [])
            guias = _so_guias(arquivos)
            log.append(f"Busca por nome: {len(arquivos)} resultado(s), {len(guias)} Guia(s).")
        except HttpError as e:
            log.append(f"Falha na busca por nome → {_msg_http(e)}")

    return guias, log


# ------------------------------------------------------------------ #
# Planilha -> JSON (mesma lógica validada do gerador local)
# ------------------------------------------------------------------ #
def _num(v, d=0):
    try:
        if v is None or v == "":
            return d
        return float(v)
    except (TypeError, ValueError):
        return d


def _parse_guia(xls_bytes, nome_arquivo, modified_iso):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xls_bytes), data_only=True)
    grupos = []
    for aba in wb.sheetnames:
        ws = wb[aba]
        for r in range(3, ws.max_row + 1):
            n = ws.cell(r, 2).value
            if n is None:
                continue
            if str(ws.cell(r, 3).value or "").strip().lower() != "andamento":
                continue
            vagas = int(_num(ws.cell(r, 26).value))
            valores = []
            for x in str(ws.cell(r, 17).value or "").split(","):
                x = x.strip()
                if x:
                    try:
                        valores.append(int(round(float(x))))
                    except ValueError:
                        pass
            valores = sorted(set(valores))
            if not valores:
                continue
            e_raw = str(ws.cell(r, 5).value or "").strip().replace(",", ".")
            try:
                fixo_pct = float(e_raw)
                fixo_ok = fixo_pct > 0
            except ValueError:
                fixo_ok = False
                fixo_pct = None
            grupos.append({
                "grupo": int(n), "aba": aba,
                "prazo": int(_num(ws.cell(r, 11).value)),
                "fundo": _num(ws.cell(r, 14).value),
                "vagas": vagas,
                "fixoOk": fixo_ok, "fixoPct": fixo_pct,
                # Coluna D (Parcela Reduzida): "S" = o grupo aceita parcela reduzida
                "red": str(ws.cell(r, 4).value or "").strip().upper() == "S",
                "valores": valores,
                "lances": [_num(ws.cell(r, 18).value), _num(ws.cell(r, 19).value),
                           _num(ws.cell(r, 22).value), _num(ws.cell(r, 23).value)],
                "contLivre": int(_num(ws.cell(r, 20).value)) + int(_num(ws.cell(r, 24).value)),
            })

    dt = _data_do_nome(nome_arquivo)
    if dt is None and modified_iso:
        try:
            dt = datetime.datetime.fromisoformat(modified_iso.replace("Z", "+00:00")).date()
        except Exception:
            dt = datetime.date.today()
    baixado = (dt or datetime.date.today()).strftime("%d/%m/%y")

    return {
        "meta": {
            "gerado_em": datetime.date.today().isoformat(),
            "fonte": nome_arquivo,
            "baixado_em": baixado,
            "total": len(grupos),
        },
        "grupos": grupos,
    }


# ------------------------------------------------------------------ #
# Carga (cache 5 min) — levanta erro claro em vez de quebrar feio
# ------------------------------------------------------------------ #
@st.cache_data(ttl=300, show_spinner=False)
def _carregar_dados_itau():
    service = get_drive_service()
    if not service:
        raise RuntimeError("Service account do Google Drive não configurado (secret `gcp_service_account`).")

    folder_id = st.secrets.get("DRIVE_FOLDER_ITAU")
    guias, log = _buscar_guias(service, folder_id)

    if not guias:
        raise RuntimeError("Nenhuma Guia de Oportunidades encontrada.\n\n" + "\n".join(f"• {l}" for l in log))

    guias.sort(key=lambda a: (_data_do_nome(a["name"]) or datetime.date.min, a.get("modifiedTime", "")))
    guia = guias[-1]

    try:
        conteudo = service.files().get_media(fileId=guia["id"]).execute()
    except HttpError as e:
        raise RuntimeError(f"Não consegui baixar '{guia['name']}' → {_msg_http(e)}")

    dados = _parse_guia(conteudo, guia["name"], guia.get("modifiedTime"))
    log.append(f"Guia escolhida: {guia['name']} ({dados['meta']['total']} grupos).")
    return dados, log


# ------------------------------------------------------------------ #
# Render
# ------------------------------------------------------------------ #
def render_itau_v2(pasta_atual):
    """Renderiza o Simulador Itaú V 2.0 com os dados vivos da Guia (Google Drive)."""
    col1, col2 = st.columns([6, 1])
    with col2:
        if st.button("🔄 Recarregar Guia", use_container_width=True):
            _carregar_dados_itau.clear()
            st.rerun()

    try:
        with st.spinner("Carregando a Guia de Oportunidades mais recente…"):
            dados, log = _carregar_dados_itau()
    except Exception as e:
        st.error("⚠️ Não foi possível carregar a Guia de Oportunidades do Google Drive.")
        st.code(str(e), language="text")
        try:
            email = dict(st.secrets["gcp_service_account"]).get("client_email", "(não encontrado)")
        except Exception:
            email = "(secret gcp_service_account ausente)"
        st.markdown(f"""
- **Service account:** `{email}`
- **DRIVE_FOLDER_ITAU:** `{st.secrets.get("DRIVE_FOLDER_ITAU", "(não definido)")}`
""")
        with st.expander("🔎 O que o service account enxerga hoje (clique para diagnosticar)", expanded=True):
            svc = get_drive_service()
            st.code(_diagnostico_visibilidade(svc) if svc else "Service account não configurado.", language="text")
            st.markdown(f"""
**Se a lista acima estiver vazia ou sem a pasta ITAU**, o compartilhamento não foi efetivado. Faça:
1. Google Drive → botão direito na pasta **ITAU** (ou na pasta pai **Tabelas**) → **Compartilhar**.
2. Cole `{email}` no campo de pessoas.
3. Permissão **Leitor** → **desmarque** "Notificar pessoas" → clique em **Compartilhar/Enviar**
   (⚠️ se não clicar no botão final, não salva!).
4. Reabra esta tela e clique em **🔄 Recarregar Guia**.
""")
        return

    caminho = os.path.join(pasta_atual, "itau_v2.html")
    try:
        with open(caminho, "r", encoding="utf-8") as f:
            html = f.read()
    except FileNotFoundError:
        st.error("⚠️ O arquivo `itau_v2.html` não foi encontrado no servidor (confira se está no GitHub).")
        return

    html = html.replace("__DADOS__", json.dumps(dados, ensure_ascii=False))
    components.html(html, height=1300, scrolling=True)

    with st.expander("ℹ️ Origem dos dados"):
        st.caption("\n".join(f"• {l}" for l in log))
